#!/usr/bin/env python3
"""批量从xlsx数据生成account markdown文件 - v2 with pinyin slugs and better EN"""
import openpyxl
import os
import re
import json
import hashlib
from pypinyin import lazy_pinyin
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ACCT_DIR = os.path.join(ROOT, "src", "content", "accounts")
XLSX_DIR = "/tmp/cnb-api"

# Load category mapping
with open(os.path.join(ROOT, 'zh_to_slug.json'), 'r', encoding='utf-8') as f:
    zh_to_slug = json.load(f)

# Platform name mapping
platform_zh_to_slug = {
    'B站': 'bilibili',
    '微博': 'weibo',
    '今日头条': 'toutiao',
    '百家号': 'baijiahao',
    '小红书': 'xiaohongshu',
    '快手': 'kuaishou',
}

platform_slug_to_zh = {
    'bilibili': 'B站（Bilibili）',
    'weibo': '微博',
    'toutiao': '今日头条',
    'baijiahao': '百家号',
    'xiaohongshu': '小红书',
    'kuaishou': '快手',
}

platform_slug_to_en = {
    'bilibili': 'Bilibili',
    'weibo': 'Weibo',
    'toutiao': 'Toutiao',
    'baijiahao': 'Baijiahao',
    'xiaohongshu': 'Xiaohongshu (RED)',
    'kuaishou': 'Kuaishou',
}

# Emoji mapping for avatars by category
category_emoji = {
    'agriculture': '🌾', 'humanities': '📖', 'sports': '⚽', 'health': '💊',
    'others': '📌', 'senior': '👵', 'military': '🎖️', 'animation': '🎨',
    'pets': '🐾', 'history': '🏛️', 'international': '🌍', 'entertainment': '🎬',
    'home-appliances': '🔌', 'lottery': '🎲', 'film-tv': '🎬', 'emotion': '💗',
    'comedy': '😂', 'photography': '📷', 'education': '📚', 'tech-review': '💻',
    'travel': '✈️', 'fashion': '👗', 'politics-society': '🏛️', 'automotive': '🚗',
    'law': '⚖️', 'gaming': '🎮', 'lifestyle': '☕', 'science-tech': '🔬',
    'science': '🔭', 'variety-show': '🎤', 'food': '🍜', 'career': '💼',
    'parenting': '👶', 'dance': '💃', 'finance': '💰', 'fitness': '💪',
    'music': '🎵', 'beauty': '✨',
}

# Category slug to zh/en names
category_slug_zh = {
    'agriculture': '三农', 'humanities': '人文社科', 'sports': '体育', 'health': '健康',
    'others': '其他', 'senior': '养老', 'military': '军事', 'animation': '动漫',
    'pets': '动物宠物', 'history': '历史', 'international': '国际', 'entertainment': '影视娱乐',
    'home-appliances': '家居家装', 'lottery': '彩票', 'film-tv': '影视', 'emotion': '情感',
    'comedy': '搞笑', 'photography': '摄影', 'education': '教育', 'tech-review': '科技数码',
    'travel': '旅游', 'fashion': '时尚', 'politics-society': '时政社会', 'automotive': '汽车',
    'law': '法律', 'gaming': '游戏', 'lifestyle': '生活日常', 'science-tech': '科学科技',
    'science': '科普', 'variety-show': '综艺', 'food': '美食', 'career': '职业职场',
    'parenting': '育儿', 'dance': '舞蹈', 'finance': '财经', 'fitness': '运动健身',
    'music': '音乐', 'beauty': '颜值',
}

category_slug_en = {
    'agriculture': 'Agriculture & Rural', 'humanities': 'Humanities & Social Science', 'sports': 'Sports', 'health': 'Health & Wellness',
    'others': 'Others', 'senior': 'Senior Living', 'military': 'Military & Defense', 'animation': 'Animation & ACG',
    'pets': 'Pets & Animals', 'history': 'History', 'international': 'International News', 'entertainment': 'Entertainment',
    'home-appliances': 'Home Appliances', 'lottery': 'Lottery', 'film-tv': 'Film & TV', 'emotion': 'Emotion & Relationships',
    'comedy': 'Comedy & Fun', 'photography': 'Photography', 'education': 'Education', 'tech-review': 'Tech & Digital',
    'travel': 'Travel & Tourism', 'fashion': 'Fashion & Beauty', 'politics-society': 'Politics & Society', 'automotive': 'Automotive',
    'law': 'Law & Legal', 'gaming': 'Gaming', 'lifestyle': 'Lifestyle & Vlog', 'science-tech': 'Science & Technology',
    'science': 'Science Popularization', 'variety-show': 'Variety & Talk Shows', 'food': 'Food & Dining', 'career': 'Career & Workplace',
    'parenting': 'Parenting & Family', 'dance': 'Dance', 'finance': 'Finance & Economics', 'fitness': 'Fitness & Exercise',
    'music': 'Music', 'beauty': 'Beauty & Self-Image',
}

def generate_slug(author):
    """Generate a semantic slug from author name"""
    if not author or not author.strip():
        return "unnamed"
    
    text = author.strip()
    
    # If it's pure Chinese, use pinyin
    if re.search(r'[\u4e00-\u9fff]', text):
        pinyin_list = lazy_pinyin(text)
        # Filter out non-alphanumeric
        slug = '-'.join(pinyin_list)
        slug = re.sub(r'[^a-z0-9-]', '', slug.lower())
        slug = re.sub(r'-+', '-', slug).strip('-')
        if len(slug) > 60:
            slug = slug[:60].rstrip('-')
        if slug:
            return slug
    
    # English or mixed - use existing logic
    parts = re.findall(r'[a-zA-Z0-9]+', text.lower())
    if parts:
        slug = '-'.join(parts)
        slug = re.sub(r'-+', '-', slug).strip('-')
        if len(slug) > 60:
            slug = slug[:60].rstrip('-')
        if slug:
            return slug
    
    # Fallback
    md5 = hashlib.md5(text.encode('utf-8')).hexdigest()[:8]
    return f"acct-{md5}"

def format_follower_count(count):
    """Format follower count for display"""
    if not count or count <= 0:
        return None, None
    if count >= 10000:
        zh = f"{count/10000:.1f}万".replace('.0万', '万')
    else:
        zh = str(count)
    if count >= 1000000:
        en = f"{count/1000000:.2f}M".replace('.00M', 'M')
    elif count >= 1000:
        en = f"{count/1000:.1f}K".replace('.0K', 'K')
    else:
        en = str(count)
    return zh, en

def get_content_styles(platform):
    """Get content styles based on platform"""
    styles = {
        'bilibili': ['中长视频', '深度内容'],
        'weibo': ['短视频', '图文动态'],
        'toutiao': ['图文资讯', '短视频'],
        'baijiahao': ['图文内容', '视频内容'],
        'xiaohongshu': ['图文种草', '短视频'],
        'kuaishou': ['短视频', '直播内容'],
    }
    return styles.get(platform, ['内容创作'])

def get_profile_url(platform, platform_id):
    """Generate a profile URL if not available"""
    base_urls = {
        'bilibili': 'https://space.bilibili.com',
        'weibo': 'https://weibo.com/u',
        'toutiao': 'https://www.toutiao.com/c/user',
        'baijiahao': 'https://baijiahao.baidu.com',
        'xiaohongshu': 'https://www.xiaohongshu.com/user/profile',
        'kuaishou': 'https://www.kuaishou.com/profile',
    }
    base = base_urls.get(platform, 'https://weibo.com')
    return f"{base}/{platform_id}"

def extract_platform_id(profile_url, slug):
    """Extract platform ID from profile URL"""
    if not profile_url:
        return f"{slug}-id"
    m = re.search(r'/(\d+)/?$', profile_url)
    if m:
        return m.group(1)
    if '/u/' in profile_url:
        m = re.search(r'/u/(\d+)', profile_url)
        if m:
            return m.group(1)
    if '/user/' in profile_url:
        m = re.search(r'/user/([a-f0-9]+|\d+)', profile_url)
        if m:
            return m.group(1)
    if '/c/user/' in profile_url:
        m = re.search(r'/c/user/(\d+)/', profile_url)
        if m:
            return m.group(1)
    return f"{slug}-id"

def generate_account_md(account, existing_slugs):
    """Generate markdown for a single account"""
    author = account['作者'].strip()
    platform_zh = account['平台'].strip()
    category_zh = account['分类'].strip()
    follower = account.get('粉丝数', 0) or 0
    growth = account.get('涨粉量', 0) or 0
    profile_url = account.get('主页链接', '') or ''
    
    platform = platform_zh_to_slug.get(platform_zh, 'bilibili')
    category = zh_to_slug.get(category_zh, 'others')
    
    slug = generate_slug(author)
    if not slug or slug in existing_slugs:
        base_slug = slug or "acct"
        i = 1
        while f"{base_slug}-{i}" in existing_slugs:
            i += 1
        slug = f"{base_slug}-{i}"
    
    existing_slugs.add(slug)
    
    # Format follower count
    fcz, fce = format_follower_count(follower)
    
    # Extract platform ID
    platform_id = extract_platform_id(profile_url, slug)
    
    # Avatar
    avatar = category_emoji.get(category, '📌')
    
    # Platform display names
    plat_zh = platform_slug_to_zh.get(platform, platform_zh)
    plat_en = platform_slug_to_en.get(platform, platform_zh)
    
    # Category display names
    cat_zh = category_slug_zh.get(category, category_zh)
    cat_en = category_slug_en.get(category, category_zh)
    
    # Content styles
    content_styles = get_content_styles(platform)
    
    # Monetization
    monetization = 'ads'
    if platform in ('xiaohongshu', 'kuaishou'):
        monetization = 'brand-deals'
    
    # English name: use pinyin for names with Chinese characters
    en_name = author
    if re.search(r'[\u4e00-\u9fff]', author):
        # Convert Chinese characters to pinyin, keep English/digits
        parts = lazy_pinyin(author)
        # Clean up: remove non-alphanumeric separators
        cleaned = []
        for p in parts:
            # Remove special chars like _ ☀ etc.
            p = re.sub(r'[_☀⚡★☆\-\\/]+', ' ', p)
            p = p.strip()
            if p:
                cleaned.append(p)
        en_name = ' '.join(cleaned)
        # Capitalize
        en_name = en_name.title()
        # If result is too long, keep first 3 words + ellipsis
        words = en_name.split()
        if len(words) > 8:
            en_name = ' '.join(words[:8])
    
    # Taglines
    tagline_zh = f"{plat_zh}上的{cat_zh}领域创作者"
    tagline_en = f"{plat_en} creator focused on {cat_en.lower()}"
    
    # Description
    desc_zh = f"{author} 是{plat_zh}上的{cat_zh}领域创作者，内容覆盖{cat_zh}相关主题与资讯分享。"
    desc_en = f"{en_name} is a {plat_en} creator focused on {cat_en.lower()}, sharing content and insights in this niche."
    
    if follower > 0:
        desc_zh += f" 公开资料显示粉丝约{fcz}。"
        desc_en += f" Public data shows approximately {fce} followers."
    
    # Build SEO
    seo_title_zh = f"{author} - {cat_zh}创作者 | 黎明岛"
    seo_title_en = f"{en_name} - {cat_en} Creator | Dawn Island"
    meta_desc_zh = f"{author}是{plat_zh}上的{cat_zh}创作者，关注{cat_zh}领域内容，公开资料显示粉丝约{fcz if fcz else '未知'}。"
    meta_desc_en = f"{en_name} is a {plat_en} creator in {cat_en.lower()}, with approximately {fce if fce else 'unknown'} followers."
    
    # GEO
    geo_summary_zh = f"{author} 是{plat_zh}上的{cat_zh}领域创作者，内容围绕{cat_zh}相关主题展开。"
    geo_summary_en = f"{en_name} is a {plat_en} creator specializing in {cat_en.lower()} content."
    
    if follower > 0:
        geo_summary_zh += f" 公开数据显示粉丝约{fcz}。"
        geo_summary_en += f" Public data shows approximately {fce} followers."
    
    # Facts
    facts_zh = []
    facts_en = []
    facts_zh.append(f"平台: \"{plat_zh}\"")
    facts_en.append(f"Platform: \"{plat_en}\"")
    facts_zh.append(f"领域: \"{cat_zh}\"")
    facts_en.append(f"Niche: \"{cat_en}\"")
    if follower > 0:
        facts_zh.append(f"粉丝量: \"约{fcz}（公开可查）\"")
        facts_en.append(f"Followers: \"~{fce} (publicly reported)\"")
    if growth and growth > 0:
        if growth >= 10000:
            growth_zh = f"{growth/10000:.1f}万".replace('.0万', '万')
        else:
            growth_zh = str(growth)
        facts_zh.append(f"近期涨粉: \"{growth_zh}\"")
        facts_en.append(f"Recent growth: \"+{growth:,}\"")
    
    # FAQ
    faq_zh = []
    faq_en = []
    faq_zh.append(f"这个账号主要讲什么？: \"{cat_zh}领域的内容创作与分享。\"")
    faq_en.append(f"What does this account cover?: \"{cat_en} content.\"")
    faq_zh.append(f"在哪里能看到？: \"{plat_zh}（主页见本页外链）。\"")
    faq_en.append(f"Where to follow?: \"{plat_en}.\"")
    if follower > 0:
        faq_zh.append(f"粉丝规模？: \"约{fcz}（公开可查，仅供参考）。\"")
        faq_en.append(f"Audience size?: \"~{fce} (publicly reported, for reference).\"")
    
    # Build YAML frontmatter
    lines = []
    lines.append(f"---")
    lines.append(f"slug: {slug}")
    lines.append(f"profileUrl: {profile_url if profile_url else get_profile_url(platform, platform_id)}")
    lines.append(f"avatar: \"{avatar}\"")
    lines.append(f"platform: {platform}")
    lines.append(f"platformId: '{platform_id}'")
    lines.append(f"verified: true")
    lines.append(f"categories:")
    lines.append(f"  - {category}")
    lines.append(f"tags:")
    lines.append(f"  - \"{cat_zh}\"")
    lines.append(f"  - \"{plat_zh}\"")
    lines.append(f"  - \"内容创作者\"")
    lines.append(f"contentStyle:")
    for style in content_styles[:3]:
        lines.append(f"  - {style}")
    lines.append(f"monetization: {monetization}")
    lines.append(f"featured: false")
    lines.append(f"draft: false")
    if follower > 0:
        lines.append(f"followerCount: {follower}")
    lines.append(f"contentFrequency: irregular")
    if growth and growth > 0 and follower > 0:
        lines.append(f"growthRate: {round(growth/follower*100, 2)}")
    lines.append(f"publishedAt: '2026-08-11'")
    lines.append(f"updatedAt: '2026-08-11'")
    lines.append(f"name:")
    lines.append(f"  en: \"{en_name}\"")
    lines.append(f"  zh: \"{author}\"")
    lines.append(f"tagline:")
    lines.append(f"  en: \"{tagline_en}\"")
    lines.append(f"  zh: \"{tagline_zh}\"")
    lines.append(f"description:")
    lines.append(f"  en: >-")
    lines.append(f"    {desc_en}")
    lines.append(f"  zh: >-")
    lines.append(f"    {desc_zh}")
    lines.append(f"seo:")
    lines.append(f"  primary_keyword: \"{author} {cat_zh} 创作者\"")
    lines.append(f"  secondary_keywords:")
    lines.append(f"    - \"{author}\"")
    lines.append(f"    - \"{cat_zh}\"")
    lines.append(f"    - \"{plat_zh}\"")
    lines.append(f"  search_intent: informational")
    lines.append(f"  title_zh: \"{seo_title_zh}\"")
    lines.append(f"  title_en: \"{seo_title_en}\"")
    lines.append(f"  meta_description_zh: \"{meta_desc_zh}\"")
    lines.append(f"  meta_description_en: \"{meta_desc_en}\"")
    lines.append(f"geo:")
    lines.append(f"  answer_summary_zh: >-")
    lines.append(f"    {geo_summary_zh}")
    lines.append(f"  answer_summary_en: >-")
    lines.append(f"    {geo_summary_en}")
    lines.append(f"  facts_zh:")
    for fact in facts_zh:
        lines.append(f"    - {fact}")
    lines.append(f"  facts_en:")
    for fact in facts_en:
        lines.append(f"    - {fact}")
    lines.append(f"  faq_zh:")
    for qa in faq_zh:
        lines.append(f"    - {qa}")
    lines.append(f"  faq_en:")
    for qa in faq_en:
        lines.append(f"    - {qa}")
    lines.append(f"---")
    
    # Add content body
    lines.append(f"<!-- zh -->")
    lines.append(f"## 他是谁")
    lines.append(f"{author} 是{plat_zh}上的{cat_zh}领域创作者。")
    lines.append(f"")
    lines.append(f"## 内容特点")
    lines.append(f"内容以{'、'.join(content_styles[:3])}为主，覆盖{cat_zh}相关主题，风格直观易懂。")
    lines.append(f"")
    lines.append(f"## 适合谁看")
    lines.append(f"适合对{cat_zh}感兴趣、希望了解该领域内容与观点的用户。")
    lines.append(f"")
    lines.append(f"## 从哪里开始")
    lines.append(f"建议在{plat_zh}关注{author}，从最新或热门内容切入，逐步建立对该领域的认知。")
    lines.append(f"")
    lines.append(f"## 常见问题")
    lines.append(f"**主要发布平台？** {plat_zh}。")
    lines.append(f"**粉丝规模？** 约{fcz if fcz else '未知'}（公开可查，仅供参考）。")
    lines.append(f"")
    lines.append(f"<!-- en -->")
    lines.append(f"## Who is this")
    lines.append(f"{en_name} is a {plat_en} creator focused on {cat_en.lower()} content.")
    lines.append(f"")
    lines.append(f"## What the content is like")
    lines.append(f"The content covers {cat_en.lower()} topics in a clear and accessible style, suitable for both newcomers and regular audiences.")
    lines.append(f"")
    lines.append(f"## Who it is for")
    lines.append(f"Anyone interested in {cat_en.lower()} and looking for quality content in this space.")
    lines.append(f"")
    lines.append(f"## Where to start")
    lines.append(f"Follow {en_name} on {plat_en}, begin with the latest or most popular posts, and build up your understanding.")
    lines.append(f"")
    lines.append(f"## FAQ")
    lines.append(f"**Main platform?** {plat_en}.")
    lines.append(f"**Audience size?** ~{fce if fce else 'unknown'} (publicly reported, for reference).")
    lines.append(f"")
    
    return '\n'.join(lines), slug

def main():
    """Read all xlsx files and generate account files"""
    # Load all data
    all_accounts = {}
    
    for f in sorted(os.listdir(XLSX_DIR)):
        if not f.endswith('.xlsx'):
            continue
        filepath = os.path.join(XLSX_DIR, f)
        wb = openpyxl.load_workbook(filepath)
        ws = wb['Sheet1']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[1]:
                continue
            acc = dict(zip(headers, row))
            key = f"{acc['作者']}|{acc['平台']}"
            if key not in all_accounts:
                all_accounts[key] = acc
            else:
                existing = all_accounts[key]
                # Prefer the one with more data
                existing_info = sum(1 for k, v in existing.items() if v)
                new_info = sum(1 for k, v in acc.items() if v)
                if new_info > existing_info:
                    all_accounts[key] = acc
    
    print(f"Total unique accounts: {len(all_accounts)}")
    
    # Track existing slugs from current files
    existing_slugs = set()
    for f in os.listdir(ACCT_DIR):
        if f.endswith('.md'):
            existing_slugs.add(f.replace('.md', ''))
    
    print(f"Existing account slugs: {len(existing_slugs)}")
    
    # Generate files
    generated = 0
    skipped = 0
    
    for key, acc in all_accounts.items():
        author = acc['作者'].strip()
        if not author:
            continue
        
        content, slug = generate_account_md(acc, existing_slugs)
        
        filename = f"{ACCT_DIR}/{slug}.md"
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(content)
        generated += 1
    
    print(f"Generated: {generated}")
    print(f"Skipped: {skipped}")

if __name__ == '__main__':
    main()
